# -*- coding: utf-8 -*-

from openpyxl import load_workbook

class XlsxReader():

    def __init__(self, filename):
        self._wb = load_workbook(filename)

    @property
    def sheetnames(self):
        return self._wb.sheetnames

    def get_sheet_content(self, sheetname):
        return self._wb[sheetname]

    def get_titles(self, sheetname):
        return [title.value for title in self.get_sheet_content(sheetname)["1"]]

    @property
    def titles(self):
        '''迭代取'''
        for sheetname in self.sheetnames:
            yield {sheetname:self.get_titles(sheetname)}
    
    @property
    def titles_dict(self):
        '''全量字典'''
        return [y for y in self.titles]

    def max_row(self, sheetname):
        return self.get_sheet_content(sheetname).max_row

    def max_column(self, sheetname):
        return self.get_sheet_content(sheetname).max_column

    def get_sheet_data(self, sheetname):
        '''获取某一张表的所有内容'''
        for idx in range(2, self.max_row(sheetname)+1):
            yield [content.value for content in self.get_sheet_content(sheetname)[idx]]
    
    def get_sheet_contents(self, sheetname):
        '''获取某一张表的所有内容'''
        sheet_data = self.get_sheet_data(sheetname)
        for data in sheet_data:
            yield {str(data[0]):data[1:]}
    
    def get_sheet_content_dict(self, sheetname):
        '''获取某一张表内容的全量字典'''
        return [y for y in self.get_sheet_contents(sheetname)]

    @property
    def contents(self):
        '''全文件data'''
        for sheetname in self.sheetnames:
            yield {sheetname:self.get_sheet_content_dict(sheetname)}
    
    @property
    def contents_dict(self):
        return [y for y in self.contents]

if __name__ == "__main__":
    x = XlsxReader("/data/config/发帖配额.xlsx")
    print(x.sheetnames)
    print(x.get_titles("1"))
    print(x.titles_dict)
    print(x.get_sheet_content_dict("1"))
    print(x.contents_dict)